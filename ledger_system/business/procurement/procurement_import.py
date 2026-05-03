"""采购清单导入主逻辑"""
from datetime import timedelta
from decimal import Decimal
from typing import Dict, List
import logging

from openpyxl import load_workbook
from ledger_system.data.database import get_session
from ledger_system.data.models import Ledger, LedgerProperty
from ledger_system.business.procurement.sheet_parsers import SHEET_PARSERS

logger = logging.getLogger(__name__)


class ProcurementImporter:
    """采购清单导入器"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.total_created = 0
        self.sheet_stats = {}

    def import_all(self) -> Dict:
        """导入所有Sheet"""
        self.wb = load_workbook(self.file_path, data_only=True)

        results = {}
        for sheet_name in self.wb.sheetnames:
            if sheet_name == "到货清单":
                results[sheet_name] = {"status": "skipped", "reason": "空Sheet"}
                continue

            parser_class = SHEET_PARSERS.get(sheet_name)
            if not parser_class:
                logger.warning(f"未找到解析器: {sheet_name}")
                results[sheet_name] = {"status": "skipped", "reason": "无解析器"}
                continue

            created = self._import_sheet(sheet_name, parser_class)
            results[sheet_name] = {"status": "success", "created": created}

        self.wb.close()
        return results

    def _import_sheet(self, sheet_name: str, parser_class) -> int:
        """导入单个Sheet"""
        ws = self.wb[sheet_name]
        created = 0

        # 跳过前两行(标题和表头)
        rows = list(ws.iter_rows(min_row=3, values_only=True))

        with get_session() as session:
            for row in rows:
                # 跳过空行
                if not any(cell is not None for cell in row):
                    continue

                try:
                    data = parser_class.parse_row(row)
                    self._create_ledger_with_properties(session, data)
                    created += 1
                except Exception as e:
                    logger.error(f"解析行失败 [{sheet_name}]: {e}, row={row}")

            session.commit()

        self.total_created += created
        self.sheet_stats[sheet_name] = created
        logger.info(f"导入完成 [{sheet_name}]: {created} 条")
        return created

    def _create_ledger_with_properties(self, session, data: Dict):
        """创建Ledger记录及关联属性"""
        # 创建Ledger
        ledger = Ledger(
            category=data["category"],
            name=data["name"],
            specification=data.get("specification", ""),
            unit=data.get("unit", ""),
            current_stock=Decimal(str(data.get("quantity", 0))),
            min_stock=Decimal("0"),
            inbound_status=data.get("inbound_status", "待入库"),
            planned_inbound_date=data.get("planned_inbound_date"),
        )
        session.add(ledger)
        session.flush()  # 获取ID

        # 创建属性记录
        properties = data.get("properties", {})
        for key, value in properties.items():
            if value is None or value == "":
                continue

            prop = LedgerProperty(
                ledger_id=ledger.id,
                property_key=str(key),
                property_value=str(value),
                property_category=self._get_property_category(key)
            )
            session.add(prop)

    @staticmethod
    def _get_property_category(key: str) -> str:
        """获取属性分类"""
        from ledger_system.data.models.ledger_property import PROPERTY_KEYS
        return PROPERTY_KEYS.get(key, ("", ""))[1] if key in PROPERTY_KEYS else ""

    @staticmethod
    def excel_date_to_date(excel_date):
        """Excel日期数字转为date对象"""
        if excel_date is None:
            return None
        if isinstance(excel_date, (int, float)):
            try:
                from datetime import datetime
                return (datetime(1899, 12, 30) + timedelta(days=int(excel_date))).date()
            except:
                return None
        if hasattr(excel_date, 'date'):
            return excel_date.date()
        return None


def import_procurement_list(file_path: str) -> Dict:
    """导入采购清单主函数"""
    importer = ProcurementImporter(file_path)
    return importer.import_all()