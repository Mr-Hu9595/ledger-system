"""Re-analyze all procurement records using AI semantic analysis"""
import sys
from pathlib import Path
import time

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import load_workbook
from ledger_system.data.database import get_session
from ledger_system.data.models import Ledger, LedgerProperty
from ledger_system.business.nlp.ai_service import AIService


def main():
    excel_path = "D:/工作/日常工作/台账/设备材料采购清单汇总表20260421.xlsx"

    print("Loading AI analyzer...")
    ai = AIService()

    print("Loading Excel file...")
    wb = load_workbook(excel_path, data_only=True)

    # Sheet to parser mapping
    sheet_info = {
        "01设备采购清单汇总表": {"start_row": 3, "name_col": 3, "qty_col": 5},
        "01-01流量计安装材料采购清单": {"start_row": 3, "name_col": 2, "qty_col": 6},
        "02材料采购清单汇总表": {"start_row": 3, "name_col": 3, "qty_col": 5},
        "02-01雾炮塔架材料清单": {"start_row": 3, "name_col": 2, "qty_col": 6},
        "02-02阀门采购清单": {"start_row": 3, "name_col": 2, "qty_col": 9},
        "02-03标准件采购清单": {"start_row": 3, "name_col": 2, "qty_col": 5},
        "03电气材料采购清单": {"start_row": 3, "name_col": 2, "qty_col": 4},
    }

    # Build list of all ledger IDs in order
    with get_session() as session:
        ledgers = session.query(Ledger).filter(
            Ledger.inbound_status == "待入库"
        ).order_by(Ledger.id).all()
        ledger_ids = [str(l.id) for l in ledgers]
        print(f"Found {len(ledger_ids)} pending inbound ledgers")

    wb.close()

    total_processed = 0
    total_updated = 0
    total_created = 0
    ledger_idx = 0

    for sheet_name, info in sheet_info.items():
        print(f"\n=== Processing {sheet_name} ===")

        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]

        start_row = info["start_row"]
        name_col = info["name_col"]

        for row_idx in range(start_row, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=name_col).value

            if not name or ledger_idx >= len(ledger_ids):
                continue

            # Get full row data
            row_data = {}
            headers = [cell.value for cell in ws[2]]
            for col_idx, header in enumerate(headers, 1):
                if header:
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(value, str) and value.startswith('='):
                        value = None
                    row_data[header] = value

            ledger_id = ledger_ids[ledger_idx]
            print(f"\n[{ledger_idx+1}/{len(ledger_ids)}] 处理: {name}")

            # Analyze with AI
            ai_result = ai.analyze_procurement_row(row_data)

            if not ai_result:
                print(f"  AI returned empty result")
                ledger_idx += 1
                continue

            # Update database with AI results
            with get_session() as session:
                updated, created = 0, 0

                for key, value in ai_result.items():
                    if value is None or value == "":
                        continue

                    existing_prop = session.query(LedgerProperty).filter(
                        LedgerProperty.ledger_id == ledger_id,
                        LedgerProperty.property_key == key
                    ).first()

                    if existing_prop:
                        if existing_prop.property_value != str(value):
                            print(f"    更新 {key}: {str(value)[:50]}...")
                            existing_prop.property_value = str(value)
                            updated += 1
                    else:
                        print(f"    新增 {key}: {str(value)[:50]}...")
                        new_prop = LedgerProperty(
                            ledger_id=ledger_id,
                            property_key=str(key),
                            property_value=str(value),
                            property_category=""
                        )
                        session.add(new_prop)
                        created += 1

                session.commit()
                total_updated += updated
                total_created += created

            total_processed += 1
            ledger_idx += 1

            # Rate limit to avoid API overload
            time.sleep(0.5)

        wb.close()

    print(f"\n=== 完成 ===")
    print(f"处理: {total_processed} 条记录")
    print(f"更新: {total_updated} 条属性")
    print(f"新增: {total_created} 条属性")


if __name__ == "__main__":
    main()