"""Material Dashboard Generator - Interactive Excel Dashboard"""
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ledger_system.data.database import get_session
from ledger_system.data.repository import LedgerRepository


class DashboardGenerator:
    """Generate interactive Excel dashboard for material lookup"""

    def __init__(self):
        self.repo = None

    def generate(self, output_path: str) -> str:
        """Generate interactive dashboard Excel file"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Sheet 1: 材料看板 (Dashboard)
        self._create_dashboard_sheet(wb)

        # Sheet 2: 台账总览 (数据源，含搜索关键字列)
        self._create_ledger_data_sheet(wb)

        # Sheet 3: 入库记录 (数据源)
        self._create_inbound_data_sheet(wb)

        # Sheet 4: 出库记录 (数据源)
        self._create_outbound_data_sheet(wb)

        wb.save(str(output_path))
        return str(output_path)

    def _create_dashboard_sheet(self, wb: Workbook) -> None:
        """Create the interactive dashboard sheet per spec 5.2"""
        ws = wb.active
        ws.title = "材料看板"

        # Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_btn_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # ===== Row 1: Title =====
        ws.merge_cells("A1:F1")
        ws["A1"] = "材料信息看板"
        ws["A1"].font = Font(bold=True, size=20)
        ws["A1"].alignment = Alignment(horizontal="center")

        # ===== Row 3: Search area =====
        ws["A3"] = "搜索材料:"
        ws["A3"].font = Font(bold=True)
        ws["B3"] = ""
        ws["B3"].fill = yellow_fill
        ws["B3"].border = thin_border
        ws["C3"] = "[搜索材料 ▼]"
        ws["C3"].font = Font(color="0070C0")
        ws["D3"] = ""
        ws["E3"] = "[+ 添加]"
        ws["E3"].font = Font(color="0070C0")

        # ===== Row 5: 已选中材料 =====
        ws.merge_cells("A5:F5")
        ws["A5"] = "已选中材料:"
        ws["A5"].fill = section_fill
        ws["A5"].font = section_font

        # Row 6: checkbox list
        ws.merge_cells("A6:F6")
        ws["A6"] = "☐（在搜索框输入关键词后，相关物料会显示在此处，选中后可添加到列表）"
        ws["A6"].font = Font(italic=True, color="888888", size=10)

        # ===== Row 8: 基本信息 =====
        ws.merge_cells("A8:F8")
        ws["A8"] = "═══ 材料基本信息 ═══"
        ws["A8"].fill = section_fill
        ws["A8"].font = section_font
        ws["A8"].alignment = Alignment(horizontal="center")

        # Row 9: 名称:  分类:  单位: labels (B9/D9/F9)
        ws["A9"] = "名称:"
        ws["A9"].font = Font(bold=True)
        ws["B9"].border = thin_border
        ws["C9"] = "分类:"
        ws["C9"].font = Font(bold=True)
        ws["D9"].border = thin_border
        ws["E9"] = "单位:"
        ws["E9"].font = Font(bold=True)
        ws["F9"].border = thin_border

        # Row 10: 规格:  物料编码:  采购日期: labels and values
        ws["A10"] = "规格:"
        ws["A10"].font = Font(bold=True)
        ws["B10"].border = thin_border
        ws["C10"] = "物料编码:"
        ws["C10"].font = Font(bold=True)
        ws["D10"].border = thin_border
        ws["E10"] = "采购日期:"
        ws["E10"].font = Font(bold=True)
        ws["F10"].border = thin_border

        # ===== Row 12: 物料属性 =====
        ws.merge_cells("A12:F12")
        ws["A12"] = "═══ 物料属性 ═══"
        ws["A12"].fill = section_fill
        ws["A12"].font = section_font
        ws["A12"].alignment = Alignment(horizontal="center")

        # Row 13: 执行标准:  材质:  介质: (B13/D13/F13)
        ws["A13"] = "执行标准:"
        ws["A13"].font = Font(bold=True)
        ws["B13"].border = thin_border
        ws["C13"] = "材质:"
        ws["C13"].font = Font(bold=True)
        ws["D13"].border = thin_border
        ws["E13"] = "介质:"
        ws["E13"].font = Font(bold=True)
        ws["F13"].border = thin_border

        # ===== Row 15: 库存情况 =====
        ws.merge_cells("A15:F15")
        ws["A15"] = "═══ 库存情况 ═══"
        ws["A15"].fill = section_fill
        ws["A15"].font = section_font
        ws["A15"].alignment = Alignment(horizontal="center")

        # Row 16: 4 column headers
        stock_headers = ["当前库存", "最小库存", "计划采购", "库存状态"]
        stock_cols = ["A", "B", "C", "D"]
        for col, header in zip(stock_cols, stock_headers):
            ws[f"{col}16"] = header
            ws[f"{col}16"].font = header_font
            ws[f"{col}16"].fill = header_fill
            ws[f"{col}16"].border = thin_border
            ws[f"{col}16"].alignment = Alignment(horizontal="center")

        # Row 17: values
        for col in stock_cols:
            ws[f"{col}17"].border = thin_border
        ws["A17"] = '=IFERROR(INDEX(台账总览!F:F,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["B17"] = '=IFERROR(INDEX(台账总览!G:G,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["C17"] = '=IFERROR(INDEX(台账总览!H:H,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["D17"] = '=IF(B3="","",IF(A17>=B17,"✓ 正常","⚠️ 库存不足"))'

        # ===== Row 19: 入库情况 =====
        ws.merge_cells("A19:F19")
        ws["A19"] = "═══ 入库情况 ═══"
        ws["A19"].fill = section_fill
        ws["A19"].font = section_font
        ws["A19"].alignment = Alignment(horizontal="center")

        # Row 20: 累计入库:  X吨  第X次入库  最近: 日期 供应商 X吨
        ws["A20"] = "累计入库:"
        ws["A20"].font = Font(bold=True)
        ws["B20"] = '=IFERROR(INDEX(台账总览!I:I,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["C20"] = "吨  第"
        ws["D20"] = '=IFERROR(INDEX(台账总览!H:H,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["D20"].font = Font(bold=True)
        ws["E20"] = "次入库  最近:"

        # Row 21: 入库 table header
        ib_headers = ["序号", "日期", "数量", "供应商", "操作人"]
        ib_cols = ["A", "B", "C", "D", "E"]
        for col, header in zip(ib_cols, ib_headers):
            ws[f"{col}21"] = header
            ws[f"{col}21"].font = header_font
            ws[f"{col}21"].fill = header_fill
            ws[f"{col}21"].border = thin_border
            ws[f"{col}21"].alignment = Alignment(horizontal="center")

        # Rows 22-26: inbound data rows
        for row in range(22, 27):
            ws.cell(row=row, column=1, value=f"第{row-21}次").border = thin_border
            for col in range(2, 6):
                ws.cell(row=row, column=col).border = thin_border

        # ===== Row 28: 出库情况 =====
        ws.merge_cells("A28:F28")
        ws["A28"] = "═══ 出库情况 ═══"
        ws["A28"].fill = section_fill
        ws["A28"].font = section_font
        ws["A28"].alignment = Alignment(horizontal="center")

        # Row 29: 累计出库:  X吨  第X次出库  最近: 日期 用途 X吨
        ws["A29"] = "累计出库:"
        ws["A29"].font = Font(bold=True)
        ws["B29"] = '=IFERROR(INDEX(台账总览!K:K,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["C29"] = "吨  第"
        ws["D29"] = '=IFERROR(INDEX(台账总览!J:J,SUMPRODUCT(MAX((ISNUMBER(SEARCH($B$3,台账总览!A:A)))*ROW(台账总览!A:A))))),"")'
        ws["D29"].font = Font(bold=True)
        ws["E29"] = "次出库  最近:"

        # Row 30: 出库 table header
        ob_headers = ["序号", "日期", "数量", "用途", "领用人"]
        ob_cols = ["A", "B", "C", "D", "E"]
        for col, header in zip(ob_cols, ob_headers):
            ws[f"{col}30"] = header
            ws[f"{col}30"].font = header_font
            ws[f"{col}30"].fill = header_fill
            ws[f"{col}30"].border = thin_border
            ws[f"{col}30"].alignment = Alignment(horizontal="center")

        # Rows 31-35: outbound data rows
        for row in range(31, 36):
            ws.cell(row=row, column=1, value=f"第{row-30}次").border = thin_border
            for col in range(2, 6):
                ws.cell(row=row, column=col).border = thin_border

        # ===== Row 37: Action buttons =====
        ws["A37"] = "刷新数据"
        ws["A37"].font = Font(bold=True, color="FFFFFF")
        ws["A37"].fill = blue_btn_fill
        ws["A37"].alignment = Alignment(horizontal="center")
        ws["A37"].border = thin_border

        ws["C37"] = "查询选中条目详情"
        ws["C37"].font = Font(bold=True, color="FFFFFF")
        ws["C37"].fill = blue_btn_fill
        ws["C37"].alignment = Alignment(horizontal="center")
        ws["C37"].border = thin_border

        ws["E37"] = "合并导出选中条目"
        ws["E37"].font = Font(bold=True, color="FFFFFF")
        ws["E37"].fill = blue_btn_fill
        ws["E37"].alignment = Alignment(horizontal="center")
        ws["E37"].border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 14

    def _create_ledger_data_sheet(self, wb: Workbook) -> None:
        """Create ledger overview data sheet with search keyword column"""
        ws = wb.create_sheet("台账总览")

        with get_session() as session:
            repo = LedgerRepository(session)
            ledgers = repo.get_all_ledgers()

            # Column A: 搜索关键字 (name | spec | category | material_code)
            headers = ["搜索关键字", "名称", "规格", "类别", "单位", "当前库存", "最小库存",
                       "入库次数", "累计入库", "出库次数", "累计出库", "净入库量", "物料编码", "采购日期", "ID"]

            # Header row
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data rows
            for row_idx, l in enumerate(ledgers, 2):
                inbounds = repo.get_inbound_history(l.id, days=99999)
                outbounds = repo.get_outbound_history(l.id, days=99999)
                cumulative_in = sum(ib.quantity for ib in inbounds) if inbounds else Decimal(0)
                cumulative_out = sum(ob.quantity for ob in outbounds) if outbounds else Decimal(0)
                net_in = cumulative_in - cumulative_out
                inbound_count = len(inbounds)
                outbound_count = len(outbounds)

                # 搜索关键字: 合并名称/规格/类别/物料编码
                search_keyword = " ".join(filter(None, [
                    l.name or "",
                    l.specification or "",
                    l.category or "",
                    l.material_code or ""
                ]))

                ws.cell(row=row_idx, column=1, value=search_keyword)
                ws.cell(row=row_idx, column=2, value=l.name)
                ws.cell(row=row_idx, column=3, value=l.specification)
                ws.cell(row=row_idx, column=4, value=l.category)
                ws.cell(row=row_idx, column=5, value=l.unit)
                ws.cell(row=row_idx, column=6, value=float(l.current_stock))
                ws.cell(row=row_idx, column=7, value=float(l.min_stock))
                ws.cell(row=row_idx, column=8, value=inbound_count)
                ws.cell(row=row_idx, column=9, value=float(cumulative_in))
                ws.cell(row=row_idx, column=10, value=outbound_count)
                ws.cell(row=row_idx, column=11, value=float(cumulative_out))
                ws.cell(row=row_idx, column=12, value=float(net_in))
                ws.cell(row=row_idx, column=13, value=l.material_code or "")
                ws.cell(row=row_idx, column=14, value=l.purchase_date.isoformat() if l.purchase_date else "")
                ws.cell(row=row_idx, column=15, value=str(l.id))

        ws.freeze_panes = "A2"

    def _create_inbound_data_sheet(self, wb: Workbook) -> None:
        """Create inbound records data sheet"""
        ws = wb.create_sheet("入库记录数据")

        with get_session() as session:
            repo = LedgerRepository(session)
            inbounds = repo.get_all_inbounds(days=99999)

            headers = ["物料名称", "序号", "日期", "时间", "数量", "单位",
                      "累计入库", "供应商", "操作人", "备注"]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row_idx, ib in enumerate(inbounds, 2):
                ledger = repo.get_ledger_by_id(ib.ledger_id)
                ws.cell(row=row_idx, column=1, value=ledger.name if ledger else "")
                ws.cell(row=row_idx, column=2, value=ib.inbound_sequence)
                ws.cell(row=row_idx, column=3, value=ib.inbound_date.isoformat())
                ws.cell(row=row_idx, column=4, value=str(ib.inbound_time) if ib.inbound_time else "")
                ws.cell(row=row_idx, column=5, value=float(ib.quantity))
                ws.cell(row=row_idx, column=6, value=ledger.unit if ledger else "")
                ws.cell(row=row_idx, column=7, value=float(ib.cumulative_in))
                ws.cell(row=row_idx, column=8, value=ib.supplier)
                ws.cell(row=row_idx, column=9, value=ib.inbound_operator)
                ws.cell(row=row_idx, column=10, value=ib.notes)

        ws.freeze_panes = "A2"

    def _create_outbound_data_sheet(self, wb: Workbook) -> None:
        """Create outbound records data sheet"""
        ws = wb.create_sheet("出库记录数据")

        with get_session() as session:
            repo = LedgerRepository(session)
            outbounds = repo.get_all_outbounds(days=99999)

            headers = ["物料名称", "序号", "日期", "时间", "数量", "单位",
                      "累计出库", "用途", "领用人", "操作人", "备注"]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for row_idx, ob in enumerate(outbounds, 2):
                ledger = repo.get_ledger_by_id(ob.ledger_id)
                ws.cell(row=row_idx, column=1, value=ledger.name if ledger else "")
                ws.cell(row=row_idx, column=2, value=ob.outbound_sequence)
                ws.cell(row=row_idx, column=3, value=ob.outbound_date.isoformat())
                ws.cell(row=row_idx, column=4, value=str(ob.outbound_time) if ob.outbound_time else "")
                ws.cell(row=row_idx, column=5, value=float(ob.quantity))
                ws.cell(row=row_idx, column=6, value=ledger.unit if ledger else "")
                ws.cell(row=row_idx, column=7, value=float(ob.cumulative_out))
                ws.cell(row=row_idx, column=8, value=ob.usage)
                ws.cell(row=row_idx, column=9, value=ob.receiver)
                ws.cell(row=row_idx, column=10, value=ob.outbound_operator)
                ws.cell(row=row_idx, column=11, value=ob.notes)

        ws.freeze_panes = "A2"
