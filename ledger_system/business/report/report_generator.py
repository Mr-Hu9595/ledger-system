"""Excel Report Generator for Ledger System"""
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Optional
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from ledger_system.data.database import get_session
from ledger_system.data.repository import LedgerRepository


class ReportGenerator:
    """Generate Excel reports from ledger data"""

    def __init__(self):
        self.repo = None

    def _get_repo(self):
        if self.repo is None:
            session = get_session().__enter__()
            self.repo = LedgerRepository(session)
        return self.repo

    def generate_full_report(self, output_path: str, start_date: date = None,
                            end_date: date = None, include_dashboard: bool = False) -> str:
        """Generate full report with all sheets"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with get_session() as session:
            repo = LedgerRepository(session)
            self.repo = repo

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Sheet 2: 台账总览
                self._write_ledger_overview(repo, writer)

                # Sheet 3: 入库记录
                self._write_inbound_records(repo, writer, start_date, end_date)

                # Sheet 4: 出库记录
                self._write_outbound_records(repo, writer, start_date, end_date)

                # Sheet 5: 物料编码
                self._write_material_codes(writer)

            # Add dashboard sheet at the beginning if requested
            if include_dashboard:
                self._add_dashboard_sheet(output_path)

            return str(output_path)

    def generate_selected_report(self, ledger_ids: List[str], output_path: str) -> str:
        """Generate report for selected ledger entries"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with get_session() as session:
            repo = LedgerRepository(session)
            self.repo = repo

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Sheet 1: 选中物料概览
                self._write_selected_overview(repo, ledger_ids, writer)

                # Sheet 2: 台账总览（选中）
                self._write_selected_ledger(repo, ledger_ids, writer)

                # Sheet 3: 入库记录（选中）
                self._write_selected_inbounds(repo, ledger_ids, writer)

                # Sheet 4: 出库记录（选中）
                self._write_selected_outbounds(repo, ledger_ids, writer)

            return str(output_path)

    def _write_ledger_overview(self, repo: LedgerRepository, writer) -> None:
        """Write ledger overview sheet"""
        ledgers = repo.get_all_ledgers()
        data = []

        for l in ledgers:
            # Get cumulative totals
            inbounds = repo.get_inbound_history(l.id, days=99999)
            outbounds = repo.get_outbound_history(l.id, days=99999)

            cumulative_in = sum(ib.quantity for ib in inbounds) if inbounds else Decimal(0)
            cumulative_out = sum(ob.quantity for ob in outbounds) if outbounds else Decimal(0)

            status = "正常" if l.current_stock >= l.min_stock else "库存不足"

            # 搜索关键字：包含名称、规格、类别、物料编码等，用空格分隔
            search_keyword = " ".join(filter(None, [
                l.name or "",
                l.specification or "",
                l.category or "",
                l.material_code or ""
            ]))

            data.append({
                "搜索关键字": search_keyword,
                "名称": l.name,
                "规格": l.specification,
                "类别": l.category,
                "单位": l.unit,
                "当前库存": float(l.current_stock),
                "最小库存": float(l.min_stock),
                "入库次数": len(inbounds),
                "入库累计": float(cumulative_in),
                "出库次数": len(outbounds),
                "出库累计": float(cumulative_out),
                "采购日期": l.purchase_date.isoformat() if l.purchase_date else "",
                "物料编码": l.material_code or "",
                "状态": status
            })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="台账总览", index=False)
        self._apply_styles(writer.sheets["台账总览"])

    def _write_inbound_records(self, repo: LedgerRepository, writer,
                               start_date: date, end_date: date) -> None:
        """Write inbound records sheet"""
        inbounds = repo.get_all_inbounds(days=99999)

        data = []
        for ib in inbounds:
            if start_date and ib.inbound_date < start_date:
                continue
            if end_date and ib.inbound_date > end_date:
                continue

            ledger = repo.get_ledger_by_id(ib.ledger_id)
            data.append({
                "序号": f"第{ib.inbound_sequence}次",
                "入库日期": ib.inbound_date.isoformat(),
                "入库时间": ib.inbound_time.isoformat() if ib.inbound_time else "",
                "物料名称": ledger.name if ledger else "",
                "规格": ledger.specification if ledger else "",
                "数量": float(ib.quantity),
                "单位": ledger.unit if ledger else "",
                "累计入库": float(ib.cumulative_in),
                "供应商": ib.supplier,
                "操作人": ib.inbound_operator,
                "单据来源": ib.document_source,
                "备注": ib.notes
            })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="入库记录", index=False)
        self._apply_styles(writer.sheets["入库记录"])

    def _write_outbound_records(self, repo: LedgerRepository, writer,
                               start_date: date, end_date: date) -> None:
        """Write outbound records sheet"""
        outbounds = repo.get_all_outbounds(days=99999)

        data = []
        for ob in outbounds:
            if start_date and ob.outbound_date < start_date:
                continue
            if end_date and ob.outbound_date > end_date:
                continue

            ledger = repo.get_ledger_by_id(ob.ledger_id)
            data.append({
                "序号": f"第{ob.outbound_sequence}次",
                "出库日期": ob.outbound_date.isoformat(),
                "出库时间": ob.outbound_time.isoformat() if ob.outbound_time else "",
                "物料名称": ledger.name if ledger else "",
                "规格": ledger.specification if ledger else "",
                "数量": float(ob.quantity),
                "单位": ledger.unit if ledger else "",
                "累计出库": float(ob.cumulative_out),
                "用途": ob.usage,
                "领用人": ob.receiver,
                "操作人": ob.outbound_operator,
                "备注": ob.notes
            })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="出库记录", index=False)
        self._apply_styles(writer.sheets["出库记录"])

    def _write_material_codes(self, writer) -> None:
        """Write material codes sheet"""
        from ledger_system.data.models import MaterialCode
        from ledger_system.data.database import get_session

        with get_session() as session:
            codes = session.query(MaterialCode).order_by(MaterialCode.code).all()

            data = [{
                "物料编码": c.code,
                "名称": c.name,
                "规格": c.specification,
                "单位": c.unit
            } for c in codes]

            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name="物料编码", index=False)
            self._apply_styles(writer.sheets["物料编码"])

    def _write_selected_overview(self, repo: LedgerRepository,
                                 ledger_ids: List[str], writer) -> None:
        """Write overview for selected ledger entries"""
        data = []

        for ledger_id in ledger_ids:
            from uuid import UUID
            try:
                l = repo.get_ledger_by_id(UUID(ledger_id))
            except:
                continue

            if not l:
                continue

            props = repo.get_ledger_with_properties(l.id)
            prop_dict = props.get("properties", {})

            data.append({
                "名称": l.name,
                "规格": l.specification,
                "类别": l.category,
                "单位": l.unit,
                "当前库存": float(l.current_stock),
                "物料编码": l.material_code or "",
                **prop_dict
            })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="选中物料概览", index=False)
        self._apply_styles(writer.sheets["选中物料概览"])

    def _write_selected_ledger(self, repo: LedgerRepository,
                               ledger_ids: List[str], writer) -> None:
        """Write ledger for selected entries"""
        self._write_ledger_overview_filtered(repo, ledger_ids, writer, "台账总览")

    def _write_selected_inbounds(self, repo: LedgerRepository,
                                ledger_ids: List[str], writer) -> None:
        """Write inbound records for selected entries"""
        self._write_inbound_records_filtered(repo, ledger_ids, writer, "入库记录")

    def _write_selected_outbounds(self, repo: LedgerRepository,
                                 ledger_ids: List[str], writer) -> None:
        """Write outbound records for selected entries"""
        self._write_outbound_records_filtered(repo, ledger_ids, writer, "出库记录")

    def _write_ledger_overview_filtered(self, repo, ledger_ids, writer, sheet_name):
        """Write filtered ledger overview"""
        data = []
        for ledger_id in ledger_ids:
            from uuid import UUID
            try:
                l = repo.get_ledger_by_id(UUID(ledger_id))
            except:
                continue
            if not l:
                continue

            status = "正常" if l.current_stock >= l.min_stock else "库存不足"
            data.append({
                "名称": l.name,
                "规格": l.specification,
                "类别": l.category,
                "单位": l.unit,
                "当前库存": float(l.current_stock),
                "最小库存": float(l.min_stock),
                "状态": status
            })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        self._apply_styles(writer.sheets[sheet_name])

    def _write_inbound_records_filtered(self, repo, ledger_ids, writer, sheet_name):
        """Write filtered inbound records"""
        data = []
        for ledger_id in ledger_ids:
            from uuid import UUID
            try:
                inbounds = repo.get_inbound_history(UUID(ledger_id), days=99999)
            except:
                continue

            for ib in inbounds:
                data.append({
                    "序号": f"第{ib.inbound_sequence}次",
                    "入库日期": ib.inbound_date.isoformat(),
                    "数量": float(ib.quantity),
                    "累计入库": float(ib.cumulative_in),
                    "供应商": ib.supplier
                })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        self._apply_styles(writer.sheets[sheet_name])

    def _write_outbound_records_filtered(self, repo, ledger_ids, writer, sheet_name):
        """Write filtered outbound records"""
        data = []
        for ledger_id in ledger_ids:
            from uuid import UUID
            try:
                outbounds = repo.get_outbound_history(UUID(ledger_id), days=99999)
            except:
                continue

            for ob in outbounds:
                data.append({
                    "序号": f"第{ob.outbound_sequence}次",
                    "出库日期": ob.outbound_date.isoformat(),
                    "数量": float(ob.quantity),
                    "累计出库": float(ob.cumulative_out),
                    "用途": ob.usage
                })

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        self._apply_styles(writer.sheets[sheet_name])

    def _apply_styles(self, sheet):
        """Apply standard styles to sheet"""
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    def _add_dashboard_sheet(self, output_path: Path) -> None:
        """Add dashboard sheet at the beginning of the Excel file"""
        from openpyxl import load_workbook

        wb = load_workbook(str(output_path))

        # Create dashboard sheet at the end first
        ws = wb.create_sheet("材料看板")
        self._write_dashboard_content(ws)

        # Move dashboard sheet to the beginning (position 0)
        # sheets are ordered by their position in wb._sheets list
        dashboard_sheet = wb["材料看板"]
        wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(dashboard_sheet)))

        wb.save(str(output_path))

    def _write_dashboard_content(self, ws) -> None:
        """Write dashboard content to sheet"""
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Title
        ws.merge_cells("A1:L1")
        ws["A1"] = "材料信息看板"
        ws["A1"].font = Font(bold=True, size=18)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Row 3: Search area
        ws["A3"] = "搜索材料:"
        ws["A3"].font = Font(bold=True)
        ws["B3"] = ""  # User input cell - VLOOKUP will be used in 台账总览 sheet
        ws["B3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws["B3"].border = thin_border
        ws["C3"] = "← 输入物料名称后查看详情（需在台账总览中查找）"
        ws["C3"].font = Font(italic=True, color="666666")

        # === Section 1: Basic Info ===
        ws.merge_cells("A5:L5")
        ws["A5"] = "基本信息"
        ws["A5"].fill = section_fill
        ws["A5"].font = section_font

        # Basic info labels
        basic_info = [
            ("名称", "B6", "D6"),
            ("规格", "B7", "D7"),
            ("类别", "B8", "D8"),
            ("单位", "F6", "G6"),
            ("物料编码", "F7", "G7"),
            ("采购日期", "F8", "G8"),
            ("当前库存", "I6", "J6"),
            ("最小库存", "I7", "J7"),
            ("库存状态", "I8", "J8"),
            ("入库次数", "I9", "J9"),
            ("入库累计", "I10", "J10"),
            ("出库次数", "K9", "L9"),
            ("出库累计", "K10", "L10"),
        ]

        for label, label_cell, value_cell in basic_info:
            ws[label_cell] = label + ":"
            ws[label_cell].font = Font(bold=True)
            ws[label_cell].border = thin_border
            ws[value_cell].border = thin_border

        # 模糊匹配公式 - 使用 AGGREGATE 找到匹配行，然后 INDEX 获取值
        # AGGREGATE(15,6,...) = SMALL ignoring errors, 15 = row number function
        # 台账总览列映射: A=搜索关键字,B=名称,C=规格,D=类别,E=单位,F=当前库存,G=最小库存,H=入库次数,I=入库累计,J=出库次数,K=出库累计,L=采购日期,M=物料编码,N=状态
        search_formula = 'AGGREGATE(15,6,ROW(台账总览!$A$2:$A$1000)/(ISNUMBER(SEARCH($B$3,台账总览!$A$2:$A$1000))),1)'
        ws["D6"] = f'=IFERROR(INDEX(台账总览!$B$2:$B$1000,{search_formula}-1),"")'  # 名称
        ws["D7"] = f'=IFERROR(INDEX(台账总览!$C$2:$C$1000,{search_formula}-1),"")'  # 规格
        ws["D8"] = f'=IFERROR(INDEX(台账总览!$D$2:$D$1000,{search_formula}-1),"")'  # 类别
        ws["G6"] = f'=IFERROR(INDEX(台账总览!$E$2:$E$1000,{search_formula}-1),"")'  # 单位
        ws["G7"] = f'=IFERROR(INDEX(台账总览!$M$2:$M$1000,{search_formula}-1),"")'  # 物料编码
        ws["G8"] = f'=IFERROR(INDEX(台账总览!$L$2:$L$1000,{search_formula}-1),"")'  # 采购日期
        ws["J6"] = f'=IFERROR(INDEX(台账总览!$F$2:$F$1000,{search_formula}-1),"")'  # 当前库存
        ws["J7"] = f'=IFERROR(INDEX(台账总览!$G$2:$G$1000,{search_formula}-1),"")'  # 最小库存
        ws["J8"] = f'=IFERROR(INDEX(台账总览!$H$2:$H$1000,{search_formula}-1),"")'  # 入库次数
        ws["J9"] = f'=IFERROR(INDEX(台账总览!$I$2:$I$1000,{search_formula}-1),"")'  # 入库累计
        ws["K8"] = f'=IFERROR(INDEX(台账总览!$J$2:$J$1000,{search_formula}-1),"")'  # 出库次数
        ws["K9"] = f'=IFERROR(INDEX(台账总览!$K$2:$K$1000,{search_formula}-1),"")'  # 出库累计
        ws["J10"] = '=IF(B3="","",IF(J6>=J7,"✓ 正常","⚠️ 库存不足"))'  # 库存状态

        # === Section 2: Inbound History ===
        ws.merge_cells("A10:L10")
        ws["A10"] = "入库记录 (最近10条)"
        ws["A10"].fill = section_fill
        ws["A10"].font = section_font

        inbound_headers = ["序号", "日期", "时间", "数量", "单位", "累计入库", "供应商", "操作人", "备注"]
        for col, header in enumerate(inbound_headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row in range(12, 22):
            ws.cell(row=row, column=1, value=f"第{row-11}次").border = thin_border
            for col in range(2, 10):
                ws.cell(row=row, column=col).border = thin_border

        # === Section 3: Outbound History ===
        ws.merge_cells("A24:L24")
        ws["A24"] = "出库记录 (最近10条)"
        ws["A24"].fill = section_fill
        ws["A24"].font = section_font

        outbound_headers = ["序号", "日期", "时间", "数量", "单位", "累计出库", "用途", "领用人", "操作人", "备注"]
        for col, header in enumerate(outbound_headers, 1):
            cell = ws.cell(row=25, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row in range(26, 36):
            ws.cell(row=row, column=1, value=f"第{row-25}次").border = thin_border
            for col in range(2, 11):
                ws.cell(row=row, column=col).border = thin_border

        # Set column widths
        column_widths = {
            "A": 12, "B": 15, "C": 15, "D": 15, "E": 10, "F": 15,
            "G": 18, "H": 12, "I": 15, "J": 12, "K": 15, "L": 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
