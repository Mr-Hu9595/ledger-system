"""Excel report synchronization - keeps Excel in sync with database"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from ledger_system.data.database import get_session
from ledger_system.data.repository import LedgerRepository


class ReportSync:
    """同步数据库数据到Excel报表"""

    REPORT_FILE = Path("D:/工作/日常工作/台账/台账报表.xlsx")

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        self._ensure_report_file()

    def _ensure_report_file(self):
        """确保报表文件存在，创建初始结构"""
        if self.REPORT_FILE.exists():
            wb = load_workbook(self.REPORT_FILE)
            # Remove old Sheet1 if exists
            if "Sheet1" in wb.sheetnames:
                del wb["Sheet1"]
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            wb.save(self.REPORT_FILE)
            wb.close()
            return

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.create_sheet("台账总览")
        wb.create_sheet("入库记录")
        wb.create_sheet("出库记录")
        wb.create_sheet("物料编码")

        wb.save(self.REPORT_FILE)
        wb.close()

    def sync_all(self):
        """同步所有数据到Excel"""
        with get_session() as session:
            repo = LedgerRepository(session)

            self._sync_inventory(repo)
            self._sync_inbounds(repo)
            self._sync_outbounds(repo)
            self._sync_material_codes(repo)

        print(f"报表已同步: {self.REPORT_FILE}")

    def _sync_inventory(self, repo: LedgerRepository):
        """同步台账总览"""
        ledgers = repo.get_all_ledgers()
        data = []
        for item in ledgers:
            data.append({
                "物料编码": item.material_code or "",
                "名称": item.name,
                "规格": item.specification,
                "类别": item.category,
                "单位": item.unit,
                "当前库存": float(item.current_stock),
                "最小库存": float(item.min_stock),
                "驱动形式": item.drive_type or "",
                "公称直径": item.nominal_diameter or "",
                "介质": item.medium or "",
                "设计压力": item.design_pressure or "",
                "材质": item.material_type or "",
                "设计温度": item.design_temperature or "",
                "阀门位置": item.valve_position or "",
                "厂家": item.manufacturer or "",
                "采购日期": item.purchase_date or "",
                "备注": item.notes or "",
                "状态": "正常" if item.current_stock >= item.min_stock else "库存不足"
            })

        df = pd.DataFrame(data)
        self._write_to_sheet("台账总览", df)

    def _sync_inbounds(self, repo: LedgerRepository):
        """同步入库记录"""
        inbounds = repo.get_all_inbounds()
        data = []
        for ib in inbounds:
            data.append({
                "日期": ib.inbound_date,
                "时间": str(ib.inbound_time)[:8] if ib.inbound_time else "",
                "物料编码": ib.ledger.material_code if ib.ledger else "",
                "材料": ib.ledger.name if ib.ledger else "",
                "规格": ib.ledger.specification if ib.ledger else "",
                "数量": float(ib.quantity),
                "单位": ib.ledger.unit if ib.ledger else "",
                "供应商": ib.supplier,
                "入库人": ib.inbound_operator,
                "原始单据": ib.original_document_path or "",
                "单据来源": ib.document_source,
                "备注": ib.notes
            })

        df = pd.DataFrame(data)
        self._write_to_sheet("入库记录", df)

    def _sync_outbounds(self, repo: LedgerRepository):
        """同步出库记录"""
        outbounds = repo.get_all_outbounds()
        data = []
        for ob in outbounds:
            data.append({
                "日期": ob.outbound_date,
                "时间": str(ob.outbound_time)[:8] if ob.outbound_time else "",
                "物料编码": ob.ledger.material_code if ob.ledger else "",
                "材料": ob.ledger.name if ob.ledger else "",
                "规格": ob.ledger.specification if ob.ledger else "",
                "数量": float(ob.quantity),
                "单位": ob.ledger.unit if ob.ledger else "",
                "用途": ob.usage,
                "领料人": ob.receiver,
                "出库人": ob.outbound_operator,
                "原始单据": ob.original_document_path or "",
                "备注": ob.notes
            })

        df = pd.DataFrame(data)
        self._write_to_sheet("出库记录", df)

    def _sync_material_codes(self, repo: LedgerRepository):
        """同步物料编码"""
        from ledger_system.data.models.material_code import MaterialCode

        codes = repo.session.query(MaterialCode).all()
        data = []
        for mc in codes:
            data.append({
                "物料编码": mc.code,
                "名称": mc.name,
                "规格": mc.specification,
                "大类": mc.category,
                "中类": mc.mid_category,
                "小类": mc.sub_category,
                "规格码": mc.spec,
                "单位码": mc.unit,
                "供应商码": mc.supplier_code,
                "年份": mc.year,
                "流水号": mc.sequence
            })

        df = pd.DataFrame(data)
        self._write_to_sheet("物料编码", df)

    def _write_to_sheet(self, sheet_name: str, df: pd.DataFrame):
        """写入数据到指定sheet"""
        wb = load_workbook(self.REPORT_FILE)

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(self.REPORT_FILE)
        wb.close()


def sync_reports():
    """同步所有报表"""
    sync = ReportSync()
    sync.sync_all()


if __name__ == "__main__":
    sync_reports()
