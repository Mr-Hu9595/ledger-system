"""从台账报表.xlsx 同步类别到数据库"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import load_workbook
from ledger_system.data.database import get_session
from ledger_system.data.models import Ledger


def sync_categories_from_excel():
    """从Excel同步类别到数据库"""
    excel_file = Path("D:/工作/日常工作/台账/台账报表.xlsx")
    wb = load_workbook(excel_file, data_only=True)
    ws = wb['台账总览']

    # Build mapping: name -> category from Excel
    excel_categories = {}
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=2).value  # 名称
        category = ws.cell(row=row_idx, column=4).value  # 类别
        if name and category:
            excel_categories[name] = category

    wb.close()

    # Update database
    updated = 0
    skipped = 0

    with get_session() as session:
        ledgers = session.query(Ledger).all()
        for ledger in ledgers:
            if ledger.name in excel_categories:
                new_category = excel_categories[ledger.name]
                if new_category != ledger.category:
                    print(f"更新: {ledger.name} | {ledger.category} -> {new_category}")
                    ledger.category = new_category
                    updated += 1
                else:
                    skipped += 1

        session.commit()

    print(f"\n同步完成: 更新 {updated} 条, 保持 {skipped} 条")
    print(f"Excel 中共有 {len(excel_categories)} 条记录")


if __name__ == "__main__":
    sync_categories_from_excel()