from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from typing import Optional
from pathlib import Path
import base64
import io
import tempfile
import os
import html
import re

router = APIRouter()

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.docx', '.xlsx', '.xls', '.pdf'}

@router.post("/recognize")
async def recognize(
    mode: str = Form(...),  # inbound | outbound | material
    text: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None)
):
    """AI识别接口，支持文本、图片、文档"""


    try:
        content = ""

        # 1. 文本输入
        if text:
            content = text

        # 2. 文件输入
        elif file:
            filename = file.filename.lower()
            if not any(filename.endswith(ext) for ext in ALLOWED_EXTENSIONS):
                raise HTTPException(status_code=400, detail="不支持的文件类型")

            file_content = await file.read()

            # 图片文件 -> MiniMax暂不支持直接图片理解，返回提示
            if filename.endswith(('.jpg', '.jpeg', '.png', '.webp')):
                raise HTTPException(
                    status_code=400,
                    detail="MiniMax暂不支持图片直接识别。请上传Word、Excel或PDF文档，或使用文本描述。"
                )

            # Word/Excel/PDF -> 提取文本
            elif filename.endswith('.docx'):
                from docx import Document
                doc = Document(io.BytesIO(file_content))
                content = '\n'.join([p.text for p in doc.paragraphs])

            elif filename.endswith(('.xlsx', '.xls')):
                import openpyxl
                import zipfile

                # xlsx files store text in sharedStrings.xml as UTF-8
                # openpyxl's values_only doesn't decode properly due to encoding issues
                # So we extract manually from the XML
                try:
                    with zipfile.ZipFile(io.BytesIO(file_content)) as z:
                        # Step 1: Extract shared strings (they are UTF-8 encoded in the XML)
                        shared_strings = []
                        if 'xl/sharedStrings.xml' in z.namelist():
                            with z.open('xl/sharedStrings.xml') as f:
                                ss_bytes = f.read()
                                # Try UTF-8 first, if we get replacement chars try GBK
                                # Some Excel files have incorrect encoding declarations
                                try:
                                    xml_content = ss_bytes.decode('utf-8')
                                    if '�' in xml_content:
                                        # Contains replacement chars, try GBK
                                        xml_content = ss_bytes.decode('gbk', errors='replace')
                                except UnicodeDecodeError:
                                    xml_content = ss_bytes.decode('gbk', errors='replace')
                                # Extract all <t>...</t> content
                                strings_from_xml = re.findall(r'<t[^>]*>([^<]*)</t>', xml_content)
                                # Build list: each <si> element contains one or more <t> elements
                                si_matches = re.findall(r'<si>(.*?)</si>', xml_content, re.DOTALL)
                                for si in si_matches:
                                    t_matches = re.findall(r'<t[^>]*>([^<]*)</t>', si)
                                    shared_strings.append(''.join(t_matches))

                        # Step 2: Parse sheet XML to get cell values
                        sheet_data = []
                        sheet_files = [n for n in z.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
                        for sheet_file in sheet_files:
                            with z.open(sheet_file) as f:
                                sheet_xml = f.read().decode('utf-8')
                                # Parse rows
                                row_matches = re.findall(r'<row[^>]*>(.*?)</row>', sheet_xml, re.DOTALL)
                                for row_xml in row_matches:
                                    cell_matches = re.findall(r'<c r="([A-Z]+\d+)"([^>]*)>(.*?)</c>', row_xml, re.DOTALL)
                                    row_cells = {}
                                    for cell_ref, attrs, cell_content in cell_matches:
                                        col = ''.join([c for c in cell_ref if c.isalpha()])
                                        # Check if shared string (t="s")
                                        if 't="s"' in attrs:
                                            v_match = re.search(r'<v>(\d+)</v>', cell_content)
                                            if v_match and shared_strings:
                                                idx = int(v_match.group(1))
                                                row_cells[col] = html.unescape(shared_strings[idx]) if idx < len(shared_strings) else ''
                                        # Check if inline string (t="inlineStr")
                                        elif 't="inlineStr"' in attrs:
                                            # Inline strings use <is><t>...</t></is>
                                            t_match = re.search(r'<t[^>]*>([^<]*)</t>', cell_content)
                                            if t_match:
                                                row_cells[col] = html.unescape(t_match.group(1))
                                        else:
                                            # Numeric or other value
                                            v_match = re.search(r'<v>([^<]+)</v>', cell_content)
                                            if v_match:
                                                row_cells[col] = v_match.group(1)

                                    # Build row text in column order
                                    if row_cells:
                                        sorted_cols = sorted(row_cells.keys())
                                        sheet_data.append(' '.join(row_cells[c] for c in sorted_cols))

                        content = '\n'.join(sheet_data)
                except Exception as e:
                    # Fallback to basic read
                    import traceback
                    traceback.print_exc()
                    wb = openpyxl.load_workbook(io.BytesIO(file_content))
                    sheets = wb.sheetnames
                    content_parts = []
                    for sheet_name in sheets:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows(values_only=True):
                            row_text = ' '.join([str(c) if c else '' for c in row])
                            if row_text.strip():
                                content_parts.append(row_text)
                    content = '\n'.join(content_parts)

            elif filename.endswith('.pdf'):
                try:
                    import pdfplumber
                    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                        content_parts = []
                        for page in pdf.pages:
                            text = page.extract_text()
                            if text:
                                content_parts.append(text)
                        content = '\n'.join(content_parts)
                except ImportError:
                    content = "PDF提取需要安装pdfplumber库，请使用图片或Word格式"
                except Exception as e:
                    content = f"PDF提取失败: {str(e)}，请使用图片或Word格式"

        else:
            raise HTTPException(status_code=400, detail="请提供text或file参数")

        # Debug: log content length and first 200 chars
        print(f"[DEBUG] Content length: {len(content)}, first 200 chars: {repr(content[:200])}")

        # 3. 根据mode调用不同解析方法
        import json
        import requests

        # 获取API配置
        config_path = Path(__file__).parent.parent.parent.parent / "ledger_system" / "config" / "settings.yaml"
        import yaml
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        api_key = config["minimax"]["api_key"]
        api_host = config["minimax"].get("api_host", "api.minimaxi.com")
        model = config["minimax"].get("model", "MiniMax-M2.7")

        # 针对入库的优化提示词
        inbound_prompt = f"""你是一个建筑工地材料入库记录解析助手。请从以下文本中提取关键信息。

文本格式可能是：
- 表格形式：用tab或空格分隔的列
- 自然语言：如"今天进了3吨钢筋"
- 多条记录：用换行分隔的不同行

文本内容：
{content}

请严格按照以下JSON格式返回（返回记录数组，每条记录包含所有字段）：
[
  {{
    "ledger_name": "材料名称（如：扁钢、钢筋、水泥等）",
    "quantity": 数量（数字）,
    "unit": "单位（如：个、米、吨、根等）",
    "supplier": "供应商/厂家名称",
    "inbound_date": "日期（格式：YYYY-MM-DD）",
    "notes": "备注信息"
  }},
  ...
]

注意：
1. 如果表格中有"规格"列（如100*3），请忽略不要放到备注里
2. "备注"列的内容才放到notes里
3. 如果没有某字段，填写null
4. 日期格式必须是YYYY-MM-DD（如2026-04-27）
5. 返回所有找到的记录，每行/每条记录一个JSON对象
6. 只返回JSON数组，不要其他内容"""

        # 针对出库的优化提示词
        outbound_prompt = f"""你是一个建筑工地材料出库记录解析助手。请从以下文本中提取关键信息。

文本格式可能是：
- 表格形式：用tab或空格分隔的列
- 自然语言：如"今天出库5吨钢筋用于施工"
- 多条记录：用换行分隔的不同行

文本内容：
{content}

请严格按照以下JSON格式返回（返回记录数组，每条记录包含所有字段）：
[
  {{
    "ledger_name": "材料名称（如：扁钢、钢筋、水泥等）",
    "quantity": 数量（数字）,
    "unit": "单位（如：个、米、吨、根等）",
    "usage": "用途",
    "receiver": "领料人",
    "outbound_date": "日期（格式：YYYY-MM-DD）",
    "notes": "备注信息"
  }},
  ...
]

注意：
1. 如果表格中有"规格"列，请忽略不要放到备注里
2. "备注"列的内容才放到notes里
3. 如果没有某字段，填写null
4. 日期格式必须是YYYY-MM-DD
5. 返回所有找到的记录，每行/每条记录一个JSON对象
6. 只返回JSON数组，不要其他内容"""

        # 针对物料的优化提示词
        material_prompt = f"""你是一个物料信息解析助手。请从以下文本中提取关键信息。

文本内容：
{content}

请严格按照以下JSON格式返回（只返回JSON，不要其他内容）：
{{
  "name": "物料名称",
  "specification": "规格型号",
  "category": "类别",
  "unit": "单位",
  "brand": "品牌",
  "nominal_diameter": "公称直径",
  "pressure": "压力",
  "min_stock": 最小库存（数字）,
  "notes": "备注"
}}

注意：
1. 规格（如100*3）要单独提取到specification字段
2. 如果没有某字段，填写null
3. 只返回JSON格式"""

        # 选择对应的提示词
        if mode == "inbound":
            prompt = inbound_prompt
        elif mode == "outbound":
            prompt = outbound_prompt
        elif mode == "material":
            prompt = material_prompt
        else:
            raise HTTPException(status_code=400, detail="无效的mode参数")

        # 直接调用MiniMax API
        url = f"https://{api_host}/anthropic/v1/messages"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "x-api-key": api_key,
            "Content-Type": "application/json"
        }
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 16384
        }

        response = requests.post(url, headers=headers, json=payload, timeout=300)
        if response.status_code != 200:
            raise HTTPException(status_code=500, detail=f"MiniMax API error: {response.text}")

        resp_json = response.json()
        # Debug: save response to file
        try:
            with open('ai_response.json', 'w', encoding='utf-8') as f:
                f.write(str(resp_json))
        except:
            pass
        # 提取text类型的内容
        ai_text = ""
        if "content" in resp_json and len(resp_json["content"]) > 0:
            for item in resp_json["content"]:
                if item.get("type") == "text":
                    ai_text = item.get("text", "")
                    break
        try:
            with open('ai_text.txt', 'w', encoding='utf-8') as f:
                f.write(ai_text)
        except:
            pass

        # 解析JSON
        try:
            # 提取JSON数组
            start = ai_text.find("[")
            end = ai_text.rfind("]") + 1
            if start != -1 and end != 0:
                json_str = ai_text[start:end]
                result = json.loads(json_str)
            else:
                # 尝试单对象格式
                start = ai_text.find("{")
                end = ai_text.rfind("}") + 1
                if start != -1 and end != 0:
                    json_str = ai_text[start:end]
                    result = [json.loads(json_str)]
                else:
                    result = []
        except json.JSONDecodeError:
            result = []

        return {"success": True, "data": result}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))